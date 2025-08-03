from PyQt5.QtCore import (
    QSettings,
    QPropertyAnimation,
    QDate,
    QRect,
    Qt,
    QDateTime,
    QCoreApplication,
    pyqtSignal,
)

from PyQt5.QtGui import (
    QIcon,
    QIntValidator,
    QPixmap,
    QFont,
    QDragEnterEvent,
    QDropEvent,
)

from PyQt5.QtWidgets import (
    QApplication,
    QMainWindow,
    QDialog,
    QFileDialog,
    QGraphicsOpacityEffect,
    QMessageBox,
    QTableWidget,
    QCalendarWidget,
    QToolButton,
    QDateEdit,
    QStyle,
    QDateTimeEdit,
    QLayout,
    QCheckBox,
    QVBoxLayout,
    QTableWidgetItem,
    QAbstractItemView,
    QWidget,
    QListWidget,
    QListWidgetItem,
    QSizePolicy,
    QComboBox,
)

from reportlab.platypus import (
    BaseDocTemplate,
    PageTemplate,
    Frame,
    Paragraph,
    Table,
    TableStyle,
    SimpleDocTemplate,
    Image,
)

from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.units import cm
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT


from Ui_1 import Ui_MainWindow
from Ui_2 import Ui_Dialog
from Ui_3 import Ui_Calendar
import os
import shutil
import sys
import openpyxl as op
import datetime
import traceback
import re


class MainWindow(QMainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.fn = None
        self.wb = None
        self.ws = None
        self.packed_em = lambda: self.edit_main()
        self.packed_sm = lambda: self.save_main()
        self.settings = QSettings("PSBkeys", "Preferences")
        self.da = DropArea(self.ui.groupBox_6)
        self.fill_table()
        self.fill_winsize()
        self.fill_functions()
        self.fill_branches()
        self.fill_dates()
        self.fill_calendars()
        self.droparea()

    def droparea(self):
        self.da.setEnabled(False)
        self.da.setGeometry(QRect(30, 380, 131, 111))

    def fill_winsize(self):
        try:
            self.resize(self.settings.value("WindowSize"))
            self.move(self.settings.value("WindowPos"))
            max = self.settings.value("FullScreen", False, bool)
            if max:
                self.showMaximized()
        except:
            pass

    def fill_functions(self):
        self.da.selection_changed.connect(self.edit_ex)
        self.ui.comboBox_2.setInsertPolicy(QComboBox.NoInsert)
        self.ui.lineEdit_6.setValidator(QIntValidator())
        self.settings_action = self.ui.menubar.addAction("Настройки")
        self.settings_action.triggered.connect(self.open_settings)

        self.ui.pushButton_11.clicked.connect(lambda: self.reserve(self.ui.lineEdit_6.text()))
        self.ui.lineEdit_6.returnPressed.connect(lambda: self.reserve(self.ui.lineEdit_6.text()))
        self.ui.pushButton.clicked.connect(lambda: self.fill_dates(1))

        self.ui.lineEdit.returnPressed.connect(self.search)
        self.ui.lineEdit_2.returnPressed.connect(self.search)
        self.ui.lineEdit_4.returnPressed.connect(self.search)
        self.ui.lineEdit_5.returnPressed.connect(self.search)
        self.ui.comboBox_2.lineEdit().returnPressed.connect(self.search)
        self.ui.pushButton_5.clicked.connect(self.packed_em)
        self.ui.pushButton_12.clicked.connect(lambda: self.save_ex())
        self.da.file_dragged.connect(lambda: self.ui.pushButton_12.setEnabled(True))
        self.ui.pushButton_6.clicked.connect(lambda: self.proc())
        self.ui.pushButton_8.clicked.connect(lambda: self.period())

    def column_resized(self, index, old, width):
        count = 0
        for n in range(1, 12):
            if self.settings.value(f"Checkboxes/{n}", False, bool):
                count += 1
                if count == index + 1:
                    self.settings.setValue(f"Table_Width/{n}", width)

    def fill_calendars(self):
        self.calendar_3 = CustomCalendarWidget()
        self.calendar_4 = CustomCalendarWidget()
        self.ui.dateEdit_3.setCalendarWidget(self.calendar_3)
        self.ui.dateEdit_4.setCalendarWidget(self.calendar_4)
        self.calendar_3.ui.pushButton.clicked.connect(lambda: (self.fill_dates(2), self.calendar_3.parent().hide()))
        self.calendar_4.ui.pushButton.clicked.connect(lambda: (self.fill_dates(3), self.calendar_4.parent().hide()))

    def fill_dates(self, o=0):
        y = int(self.date()[2])
        m = int(self.date()[1])
        d = int(self.date()[0])
        yest_y = int(self.date(1)[2])
        yest_m = int(self.date(1)[1])
        yest_d = int(self.date(1)[0])
        if o == 0:
            self.ui.dateEdit_2.setDate(QDate(y, m, d))
            self.ui.dateEdit_3.setDate(QDate(yest_y, yest_m, yest_d))
            self.ui.dateEdit_4.setDate(QDate(y, m, d))
        elif o == 1:
            self.ui.dateEdit_2.setDate(QDate(y, m, d))
        elif o == 2:
            self.ui.dateEdit_3.setDate(QDate(y, m, d))
        elif o == 3:
            self.ui.dateEdit_4.setDate(QDate(y, m, d))

    def fill_branches(self, o=0):
        try:
            branches = self.settings.value("Path/Branches")
            names = self.settings.value("Path/Names")
            if o == 0:
                self.ui.comboBox_2.clear()
                self.ui.comboBox_5.clear()
                self.ui.comboBox_4.clear()
                self.ui.comboBox_3.clear()
                if branches:
                    b = branches.split("\n")
                    for i in b:
                        self.ui.comboBox_2.addItem(i)
                        self.ui.comboBox_5.addItem(i)
                        self.ui.comboBox_4.addItem(i)
                if names:
                    n = names.split("\n")
                    for i in n:
                        self.ui.comboBox_3.addItem(i)
            if o == 1:
                self.ui.comboBox_4.clear()
                self.ui.comboBox_3.clear()
                if branches:
                    b = branches.split("\n")
                    for i in b:
                        self.ui.comboBox_4.addItem(i)
                if names:
                    n = names.split("\n")
                    for i in n:
                        self.ui.comboBox_3.addItem(i)

        except:
            pass

    def date(self, o=0):
        today = datetime.date.today().strftime("%d.%m.%Y")
        yesterday = (datetime.date.today() - datetime.timedelta(days=1)).strftime("%d.%m.%Y")
        if o == 0:
            return today.split(".")
        else:
            return yesterday.split(".")

    def xl_name(self):
        self.fn = self.settings.value("Path/Path_1")
        self.wb = op.load_workbook(self.fn)
        self.ws = self.wb.active

    def error(self, e):
        e = str(e)
        self.msg = QMessageBox()
        self.msg.setIcon(QMessageBox.Critical)
        if not any(
            x in e
            for x in [
                "'NoneType' object has no attribute 'save'",
                "'NoneType' object is not subscriptable",
                "expected str, bytes or os.PathLike object, not NoneType",
                "'SettingsDialog' object has no attribute 'xl_name'",
                "openpyxl does not support  file format, please check you can open it with Excel first. Supported formats are: .xlsx,.xlsm,.xltx,.xltm",
            ]
        ):
            if "expected str, bytes or os.PathLike object, not NoneType" in e:
                self.msg.setText("Укажите путь к таблице в настройках")
            elif "openpyxl does not support  file format, please check you can open it with Excel first. Supported formats are: .xlsx,.xlsm,.xltx,.xltm" in e:
                self.msg.setText("Укажите путь к таблице в настройках")

            elif "Permission denied" in e:
                self.msg.setText("Закройте клиент Microsoft Excel для корректной работы")
            elif "cannot access local variable 'data' where it is not associated with a value" in e:
                self.msg.setText("Элемент не выбран")

            self.msg.setInformativeText(e)
            self.msg.setWindowTitle("Error")
            self.msg.exec_()

    def dec(f):
        print("dec")

        def wrap(self, *args, **kwargs):
            print("dec")
            main_window = self.get_mw()
            try:
                main_window.xl_name()
                f(self, *args, **kwargs)
                main_window.wb.save(main_window.fn)
                main_window.wb.close()
            except Exception as e:
                main_window.error(e)
                print(f.__name__)
                print(e)

        return wrap

    def get_mw(self):
        return self

    @dec
    def reserve(self, n):
        b = self.ui.comboBox_5.currentText()
        if n:
            n = int(n)
            last = 0
            for i in self.ws["A"]:
                if i.value:
                    if type(i.value) == int:
                        last = i.value
                else:
                    self.ws.delete_rows(i.row, 1)
            count = 0
            while count != n:
                count += 1
                self.ws.append({1: last + count, 2: b, 3: (".").join(self.date())})

            self.ui.label_4.setText(f"{n} keys added")
            self.ui.lineEdit_6.setText("")
            effect = QGraphicsOpacityEffect(self.ui.label_4)
            self.ui.label_4.setGraphicsEffect(effect)
            animation = QPropertyAnimation(effect, b"opacity", self.ui.label_4)
            animation.setDuration(1000)
            animation.setStartValue(1.0)
            animation.setEndValue(0.0)
            animation.start()
            for i in self.ws["A"]:
                print(i.value)

    @dec
    def fill_table(self):
        colcount = 0
        for i, n in zip(self.ws["1"], range(1, 12)):
            if self.settings.value(f"CheckBoxes/{n}", False, bool):
                colcount += 1
                self.ui.tableWidget.setColumnCount(colcount)
                item = QTableWidgetItem(colcount)
                self.ui.tableWidget.setHorizontalHeaderItem(colcount - 1, item)
                self.ui.tableWidget.horizontalHeaderItem(colcount - 1).setText(i.value)
                self.ui.tableWidget.horizontalHeader()
                width = self.settings.value(f"Table_Width/{n}", type=int)
                if width:
                    self.ui.tableWidget.setColumnWidth(colcount - 1, width)
                else:
                    self.ui.tableWidget.resizeColumnToContents(colcount - 1)
            else:
                self.settings.remove(f"Table_Width/{n}")

        self.ui.tableWidget.horizontalHeader().sectionResized.connect(self.column_resized)
        self.ui.tableWidget.setSelectionBehavior(QTableWidget.SelectRows)
        self.ui.tableWidget.setSelectionMode(QAbstractItemView.SingleSelection)
        self.ui.tableWidget.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.ui.tableWidget.setRowCount(0)
        self.ui.tableWidget.cellClicked.connect(lambda: self.fill_edit())

    @dec
    def search(self):
        self.ui.tableWidget.setRowCount(0)
        text_1 = self.ui.lineEdit.text()
        text_2 = self.ui.lineEdit_2.text()
        text_3 = self.ui.comboBox_2.lineEdit().text()
        text_4 = self.ui.lineEdit_4.text()
        text_5 = self.ui.lineEdit_5.text()
        filter_me = {5: text_1, 7: text_2, 1: text_3, 3: text_4, 0: text_5}
        text = [(i, b) for i, b in filter_me.items() if b]
        c = 0
        r = 0
        for row in self.ws.iter_rows(min_row=2, max_row=self.ws.max_row, min_col=1, max_col=11):
            count = 0
            for i, b in text:
                if b in str(row[i].value):
                    count += 1
                    if count == len(text):
                        r += 1
                        c = 0
                        self.ui.tableWidget.setRowCount(r)
                        for n in range(1, 12):
                            if self.settings.value(f"Checkboxes/{n}", False, bool):
                                item = QTableWidgetItem()
                                self.ui.tableWidget.setItem(r - 1, c, item)
                                item.setData(Qt.UserRole, f"{row}")
                                if row[n - 1].value:
                                    item.setText(f"{row[n-1].value}")
                                c += 1

    @dec
    def fill_edit(self):
        self.ui.lineEdit.returnPressed.connect(self.retrieve_s)
        self.ui.lineEdit_2.returnPressed.connect(self.retrieve_s)
        self.ui.lineEdit_4.returnPressed.connect(self.retrieve_s)
        self.ui.lineEdit_5.returnPressed.connect(self.retrieve_s)
        self.settings_action.triggered.connect(self.retrieve_s)
        self.ui.comboBox_2.lineEdit().returnPressed.connect(self.retrieve_s)

        self.ui.pushButton.setEnabled(True)
        self.ui.pushButton_5.setEnabled(True)
        self.ui.dateEdit_2.setEnabled(True)
        selected = self.ui.tableWidget.currentItem()
        if selected is not None:
            index = self.ui.tableWidget.indexFromItem(selected)
            data = self.ui.tableWidget.item(index.row(), index.column()).data(Qt.UserRole)
        pat = r"\.\D+\d+>"
        n = re.findall(pat, data)
        pat2 = r"[a-zA-Z]+\d+"
        res = re.findall(pat2, str(n))

        # филиал
        t = self.ws[f"{res[1]}"].value
        if t:
            self.ui.comboBox_4.lineEdit().setText(t)
        else:
            self.ui.comboBox_4.lineEdit().setText("")

        # ключи ввел
        t = self.ws[f"{res[8]}"].value
        if t:
            self.ui.comboBox_3.lineEdit().setText(str(t))
        else:
            self.ui.comboBox_3.lineEdit().setText("")

        # ид
        t = self.ws[f"{res[3]}"].value
        if t:
            self.ui.lineEdit_11.setText(str(t))
        else:
            self.ui.lineEdit_11.setText("")

        # тип пос
        t = self.ws[f"{res[4]}"].value
        if t:
            self.ui.lineEdit_12.setText(str(t))
        else:
            self.ui.lineEdit_12.setText("")

        # сер пос
        t = self.ws[f"{res[5]}"].value
        if t:
            self.ui.lineEdit_13.setText(str(t))
        else:
            self.ui.lineEdit_13.setText("")

        # тип пин
        t = self.ws[f"{res[6]}"].value
        if t:
            self.ui.lineEdit_14.setText(str(t))
        else:
            self.ui.lineEdit_14.setText("")

        # сер пин
        t = self.ws[f"{res[7]}"].value
        if t:
            self.ui.lineEdit_15.setText(str(t))
        else:
            self.ui.lineEdit_15.setText("")
        self.da.list_widget.clear()
        key = self.ws[f"{res[0]}"].value
        folder_name = key
        folder_path = self.settings.value("Path/Path_2")
        new_path = f"{folder_path}/{folder_name}"
        if os.path.exists(new_path):
            for root, dirs, files in os.walk(new_path):
                for file in files:
                    list_item = QListWidgetItem(QIcon("assets/file.svg"), file)
                    list_item.setData(Qt.UserRole, f"{new_path}/{file}")
                    self.da.list_widget.addItem(list_item)

    def edit_main(self):
        self.ui.lineEdit.returnPressed.disconnect(self.retrieve_s)
        self.ui.lineEdit_2.returnPressed.disconnect(self.retrieve_s)
        self.ui.lineEdit_4.returnPressed.disconnect(self.retrieve_s)
        self.ui.lineEdit_5.returnPressed.disconnect(self.retrieve_s)
        self.settings_action.triggered.disconnect(self.retrieve_s)
        self.ui.comboBox_2.lineEdit().returnPressed.disconnect(self.retrieve_s)

        self.ui.lineEdit.returnPressed.connect(self.retrieve_l)
        self.ui.lineEdit_2.returnPressed.connect(self.retrieve_l)
        self.ui.lineEdit_4.returnPressed.connect(self.retrieve_l)
        self.ui.lineEdit_5.returnPressed.connect(self.retrieve_l)
        self.settings_action.triggered.connect(self.retrieve_l)
        self.ui.comboBox_2.lineEdit().returnPressed.connect(self.retrieve_l)
        self.ui.tableWidget.cellClicked.connect(self.retrieve_m)

        self.ui.pushButton_5.clicked.disconnect(self.packed_em)
        self.ui.pushButton_5.clicked.connect(self.packed_sm)
        self.ui.pushButton_5.setText("Сохранить")

        # филиал
        self.ui.label_15.setEnabled(True)
        self.ui.comboBox_4.setEnabled(True)

        # ключи ввел
        self.ui.label_21.setEnabled(True)
        self.ui.comboBox_3.setEnabled(True)

        # ид
        self.ui.label_16.setEnabled(True)
        self.ui.lineEdit_11.setEnabled(True)

        # тип пос
        self.ui.label_17.setEnabled(True)
        self.ui.lineEdit_12.setEnabled(True)

        # сер пос
        self.ui.label_18.setEnabled(True)
        self.ui.lineEdit_13.setEnabled(True)

        # тип пин
        self.ui.label_19.setEnabled(True)
        self.ui.lineEdit_14.setEnabled(True)

        # сер пин
        self.ui.label_20.setEnabled(True)
        self.ui.lineEdit_15.setEnabled(True)

        # драг дроп
        self.da.setEnabled(True)
        self.ui.pushButton_12.setEnabled(True)

    @dec
    def save_main(self):
        selected = self.ui.tableWidget.currentItem()
        if selected is not None:
            index = self.ui.tableWidget.indexFromItem(selected)
            data = self.ui.tableWidget.item(index.row(), index.column()).data(Qt.UserRole)
        pat = r"\.\D+\d+>"
        n = re.findall(pat, data)
        pat2 = r"[a-zA-Z]+\d+"
        res = re.findall(pat2, str(n))
        d = {}
        count = 0
        for n in range(1, 12):
            if self.settings.value(f"CheckBoxes/{n}", False, bool):
                d[n] = count
                count += 1

        # дата ввода
        t = self.ui.dateEdit_2.text()
        self.ws[f"{res[9]}"] = t
        if 10 in list(d.keys()):
            self.ui.tableWidget.item(index.row(), d[10]).setText(t)

        # филиал
        t = self.ui.comboBox_4.lineEdit().text()
        self.ws[f"{res[1]}"] = t
        if 2 in list(d.keys()):
            self.ui.tableWidget.item(index.row(), d[2]).setText(t)

        # ключи ввел
        t = self.ui.comboBox_3.lineEdit().text()
        self.ws[f"{res[8]}"] = t
        if 9 in list(d.keys()):
            self.ui.tableWidget.item(index.row(), d[9]).setText(t)

        # ид
        t = self.ui.lineEdit_11.text()
        self.ws[f"{res[3]}"] = t
        if 4 in list(d.keys()):
            self.ui.tableWidget.item(index.row(), d[4]).setText(t)

        # тип пос
        t = self.ui.lineEdit_12.text()
        self.ws[f"{res[4]}"] = t
        if 5 in list(d.keys()):
            self.ui.tableWidget.item(index.row(), d[5]).setText(t)

        # сер пос
        t = self.ui.lineEdit_13.text()
        self.ws[f"{res[5]}"] = t
        if 6 in list(d.keys()):
            self.ui.tableWidget.item(index.row(), d[6]).setText(t)

        # тип пин
        t = self.ui.lineEdit_14.text()
        self.ws[f"{res[6]}"] = t
        if 7 in list(d.keys()):
            self.ui.tableWidget.item(index.row(), d[7]).setText(t)

        # сер пин
        t = self.ui.lineEdit_15.text()
        self.ws[f"{res[7]}"] = t
        if 8 in list(d.keys()):
            self.ui.tableWidget.item(index.row(), d[8]).setText(t)

        self.retrieve_l()
        # дроп
        self.da.list_widget.clear()
        key = self.ws[f"{res[0]}"].value
        folder_name = key
        folder_path = self.settings.value("Path/Path_2")
        new_path = f"{folder_path}/{folder_name}"
        if os.path.exists(new_path):
            for root, dirs, files in os.walk(new_path):
                for file in files:
                    list_item = QListWidgetItem(QIcon("assets/file.svg"), file)
                    list_item.setData(Qt.UserRole, f"{new_path}/{file}")
                    self.da.list_widget.addItem(list_item)

    @dec
    def save_ex(self):
        selected = self.ui.tableWidget.currentItem()
        if selected is not None:
            index = self.ui.tableWidget.indexFromItem(selected)
            data = self.ui.tableWidget.item(index.row(), index.column()).data(Qt.UserRole)
        pat = r"\.\D+\d+>"
        n = re.findall(pat, data)
        pat2 = r"[a-zA-Z]+\d+"
        res = re.findall(pat2, str(n))

        d = {}
        count = 0
        for n in range(1, 12):
            if self.settings.value(f"CheckBoxes/{n}", False, bool):
                d[n] = count
                count += 1

        # дата ввода
        t = self.ui.dateEdit_2.text()
        self.ws[f"{res[9]}"] = t
        if 10 in list(d.keys()):
            self.ui.tableWidget.item(index.row(), d[10]).setText(t)

        # филиал
        t = self.ui.comboBox_4.lineEdit().text()
        self.ws[f"{res[1]}"] = t
        if 2 in list(d.keys()):
            self.ui.tableWidget.item(index.row(), d[2]).setText(t)

        # ключи ввел
        t = self.ui.comboBox_3.lineEdit().text()
        self.ws[f"{res[8]}"] = t
        if 9 in list(d.keys()):
            self.ui.tableWidget.item(index.row(), d[9]).setText(t)

        # ид
        t = self.ui.lineEdit_11.text()
        self.ws[f"{res[3]}"] = t
        if 4 in list(d.keys()):
            self.ui.tableWidget.item(index.row(), d[4]).setText(t)

        # тип пос
        t = self.ui.lineEdit_12.text()
        self.ws[f"{res[4]}"] = t
        if 5 in list(d.keys()):
            self.ui.tableWidget.item(index.row(), d[5]).setText(t)

        # сер пос
        t = self.ui.lineEdit_13.text()
        self.ws[f"{res[5]}"] = t
        if 6 in list(d.keys()):
            self.ui.tableWidget.item(index.row(), d[6]).setText(t)

        # тип пин
        t = self.ui.lineEdit_14.text()
        self.ws[f"{res[6]}"] = t
        if 7 in list(d.keys()):
            self.ui.tableWidget.item(index.row(), d[7]).setText(t)

        # сер пин
        t = self.ui.lineEdit_15.text()
        self.ws[f"{res[7]}"] = t
        if 8 in list(d.keys()):
            self.ui.tableWidget.item(index.row(), d[8]).setText(t)

        key = str(self.ws[f"{res[0]}"].value)
        folder_name = key
        folder_path = self.settings.value("Path/Path_2")
        new_path = f"{folder_path}/{folder_name}"
        os.makedirs(new_path, exist_ok=True)
        for i in range(self.da.list_widget.count()):
            item = self.da.list_widget.item(i)
            old_path = item.data(Qt.UserRole)
            if f"{new_path}/{item.text()}" != old_path:
                if os.path.exists(f"{new_path}/{item.text()}"):
                    new_path = f"{new_path}/{os.path.splitext(item.text())[0]}_1{os.path.splitext(old_path)[1]}"
                shutil.move(old_path, new_path)
        self.retrieve_l()

    def edit_ex(self, name, data):
        print(name)
        print(data)

        self.ui.pushButton_13.setEnabled(True)

    def del_ex(self):
        # folder_path = self.settings.value("Path/Path_2")
        # new_path = f"{folder_path}/{folder_name}"
        # os.makedirs(new_path, exist_ok=True)
        # for i in range(self.da.list_widget.count()):
        #     item = self.da.list_widget.item(i)
        #     old_path = item.data(Qt.UserRole)
        #     if f"{new_path}/{item.text()}" != old_path:
        #         if os.path.exists(f"{new_path}/{item.text()}"):
        #             new_path = f"{new_path}/{os.path.splitext(item.text())[0]}_1{os.path.splitext(old_path)[1]}"
        #         shutil.move(old_path, new_path)
        pass

    def retrieve_s(self):
        self.ui.lineEdit.returnPressed.disconnect(self.retrieve_s)
        self.ui.lineEdit_2.returnPressed.disconnect(self.retrieve_s)
        self.ui.lineEdit_4.returnPressed.disconnect(self.retrieve_s)
        self.ui.lineEdit_5.returnPressed.disconnect(self.retrieve_s)
        self.settings_action.triggered.disconnect(self.retrieve_s)
        self.ui.comboBox_2.lineEdit().returnPressed.disconnect(self.retrieve_s)
        self.ui.pushButton.setEnabled(False)
        self.ui.pushButton_5.setEnabled(False)
        self.ui.dateEdit_2.setEnabled(False)

    def retrieve_m(self):
        self.ui.tableWidget.cellClicked.disconnect(self.retrieve_m)
        self.ui.pushButton_5.setText("Редактировать")
        self.ui.pushButton_5.clicked.disconnect(self.packed_sm)
        self.ui.pushButton_5.clicked.connect(self.packed_em)
        # филиал
        self.ui.label_15.setEnabled(False)
        self.ui.comboBox_4.setEnabled(False)
        # ключи ввел
        self.ui.label_21.setEnabled(False)
        self.ui.comboBox_3.setEnabled(False)
        # ид
        self.ui.label_16.setEnabled(False)
        self.ui.lineEdit_11.setEnabled(False)
        # тип пос
        self.ui.label_17.setEnabled(False)
        self.ui.lineEdit_12.setEnabled(False)
        # сер пос
        self.ui.label_18.setEnabled(False)
        self.ui.lineEdit_13.setEnabled(False)
        # тип пин
        self.ui.label_19.setEnabled(False)
        self.ui.lineEdit_14.setEnabled(False)
        # сер пин
        self.ui.label_20.setEnabled(False)
        self.ui.lineEdit_15.setEnabled(False)
        # драг дроп
        self.da.setEnabled(False)
        # сохран
        self.ui.pushButton_12.setEnabled(False)
        # удал
        self.ui.pushButton_13.setEnabled(False)

    def retrieve_l(self):
        self.ui.lineEdit.returnPressed.disconnect(self.retrieve_l)
        self.ui.lineEdit_2.returnPressed.disconnect(self.retrieve_l)
        self.ui.lineEdit_4.returnPressed.disconnect(self.retrieve_l)
        self.ui.lineEdit_5.returnPressed.disconnect(self.retrieve_l)
        self.ui.comboBox_2.lineEdit().returnPressed.disconnect(self.retrieve_l)
        self.ui.tableWidget.cellClicked.disconnect(self.retrieve_m)
        self.ui.pushButton.setEnabled(False)
        self.ui.pushButton_5.setEnabled(False)
        self.ui.dateEdit_2.setEnabled(False)
        self.ui.pushButton_5.setText("Редактировать")
        self.ui.pushButton_5.clicked.disconnect(self.packed_sm)
        self.ui.pushButton_5.clicked.connect(self.packed_em)
        # филиал
        self.ui.label_15.setEnabled(False)
        self.ui.comboBox_4.setEnabled(False)
        # ключи ввел
        self.ui.label_21.setEnabled(False)
        self.ui.comboBox_3.setEnabled(False)
        # ид
        self.ui.label_16.setEnabled(False)
        self.ui.lineEdit_11.setEnabled(False)
        # тип пос
        self.ui.label_17.setEnabled(False)
        self.ui.lineEdit_12.setEnabled(False)
        # сер пос
        self.ui.label_18.setEnabled(False)
        self.ui.lineEdit_13.setEnabled(False)
        # тип пин
        self.ui.label_19.setEnabled(False)
        self.ui.lineEdit_14.setEnabled(False)
        # сер пин
        self.ui.label_20.setEnabled(False)
        self.ui.lineEdit_15.setEnabled(False)
        # драг дроп
        self.da.setEnabled(False)
        # сохран
        self.ui.pushButton_12.setEnabled(False)
        # удал
        self.ui.pushButton_13.setEnabled(False)

    def open_settings(self):
        if not hasattr(self, "settings_dialog") or self.settings_dialog.isHidden():
            self.settings_dialog = SettingsDialog(self)
            self.settings_dialog.exec_()
        else:
            self.settings_dialog.activateWindow()

    def closeEvent(self, event):
        super().closeEvent(event)
        max = self.isMaximized()
        if max:
            self.settings.setValue("FullScreen", self.isMaximized())
        else:
            self.settings.setValue("FullScreen", self.isMaximized())
            self.settings.setValue("WindowSize", self.size())
            self.settings.setValue("WindowPos", self.pos())

    def proc(self):
        pdfmetrics.registerFont(TTFont("times", "assets/times.ttf"))
        pdfmetrics.registerFont(TTFont("timesbd", "assets/timesbd.ttf"))

        text = "Список<br/>бланков регистрации,<br/>переданных 10.10.2023 сотрудниками ОЗПК в отдел процессинга"
        text_2 = "Серым выделены ключи, загруженные с систему TWO автоматически."
        text_3 = "Записей по ключевым данным передано:<br/><br/>Ответственный сотрудник"
        text_4 = "<u>\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a00\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0</u><br/><br/><u>\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0</u>\u00a0\u00a0\u00a0\u00a0<u>/\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0 </u><br/><u>10.10.2023 11:22:05</u>"

        styles = getSampleStyleSheet()

        style1 = styles["Normal"].clone("style1")
        style1.fontName = "times"
        style1.fontSize = 14
        style1.leading = 16
        style1.alignment = TA_CENTER

        style2 = styles["Normal"].clone("style2")
        style2.fontName = "times"
        style2.fontSize = 12
        style2.leading = 16
        style2.alignment = TA_CENTER

        style3 = styles["Normal"].clone("style3")
        style3.fontName = "times"
        style3.fontSize = 12
        style3.leading = 20
        style3.alignment = TA_LEFT

        data = [
            [
                "№    ",
                "     ID Терминала     ",
                "Номер комплекта ключевых данных",
                "     Дата     ",
            ]
        ]

        table = Table(data)
        table_style = TableStyle(
            [
                (
                    "FONTNAME",
                    (0, 0),
                    (-1, 0),
                    "times",
                ),  # Задаем жирный шрифт для заголовков
                (
                    "FONTSIZE",
                    (0, 0),
                    (-1, 0),
                    14,
                ),  # Задаем размер шрифта для заголовков
                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, 0),
                    5,
                ),  # Задаем отступ сверху для заголовков
                (
                    "BACKGROUND",
                    (0, 1),
                    (-1, -1),
                    "white",
                ),  # Задаем белый фон для остальных ячеек
                ("GRID", (0, 0), (-1, -1), 0.5, "black"),  # Задаем сетку для всех ячеек
                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "MIDDLE",
                ),  # Задаем вертикальное выравнивание по середине для всех ячеек
            ]
        )
        image = Image("assets/PSB_logo.png", 6.45 * cm, 2.8 * cm)
        table.setStyle(table_style)
        p = Paragraph(text, style1)
        p_2 = Paragraph(text_2, style2)
        p_3 = Paragraph(text_3, style3)
        p_4 = Paragraph(text_4, style3)

        doc = BaseDocTemplate("Список в ПЦ.pdf", pagesize=A4)

        frame1 = Frame(0 * cm, 25.5 * cm, 8 * cm, 4 * cm, id="F1")
        frame2 = Frame(3 * cm, 24 * cm, 16 * cm, 2.5 * cm, id="F2")
        frame3 = Frame(5 * cm, 20 * cm, 12 * cm, 2 * cm, id="F3")
        frame4 = Frame(0 * cm, 19 * cm, 16 * cm, 2 * cm, id="F4")
        frame5 = Frame(1.5 * cm, 16 * cm, 20 * cm, 3 * cm, id="F5")
        frame6 = Frame(10 * cm, 15 * cm, 20 * cm, 4 * cm, id="F6")

        page_template = PageTemplate(id="PT1", frames=[frame1, frame2, frame3, frame4, frame5, frame6])

        doc.addPageTemplates([page_template])

        story = [image, p, table, p_2, p_3, p_4]

        doc.build(story)

    def period(self):
        d1 = self.ui.dateEdit_3.text()
        d2 = self.ui.dateEdit_4.text()
        pdfmetrics.registerFont(TTFont("times", "assets/times.ttf"))
        pdfmetrics.registerFont(TTFont("timesbd", "assets/timesbd.ttf"))

        text = "Начальнику отдела<br/>защиты пластиковых карт<br/>Тюрникову В.А."
        text_2 = f"Отчет<br/>о проделанной работе в период <u>'{d1}-{d2}'</u>"
        text_3 = "1. Генерация КД:<br/>\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0-для POS-терминалов - ____ шт.<br/>2. Контроль загрузки компонентов КД в СБО:<br/>\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0-для POS-терминалов - ____ шт.<br/>"

        styles = getSampleStyleSheet()

        style1 = styles["Normal"].clone("style1")
        style1.fontName = "timesbd"
        style1.fontSize = 14
        style1.leading = 16
        style1.alignment = TA_LEFT

        style2 = styles["Normal"].clone("style2")
        style2.fontName = "timesbd"
        style2.fontSize = 14
        style2.leading = 16
        style2.alignment = TA_CENTER

        style3 = styles["Normal"].clone("style3")
        style3.fontName = "times"
        style3.fontSize = 14
        style3.leading = 30
        style3.alignment = TA_LEFT

        p = Paragraph(text, style1)
        p_2 = Paragraph(text_2, style2)
        p_3 = Paragraph(text_3, style3)

        doc = BaseDocTemplate("Отчет за период.pdf", pagesize=A4)

        frame1 = Frame(12 * cm, 26 * cm, 9 * cm, 3 * cm, id="F1")
        frame2 = Frame(4 * cm, 23 * cm, 14 * cm, 2 * cm, id="F2")
        frame3 = Frame(3 * cm, 14 * cm, 12 * cm, 8 * cm, id="F3")

        page_template = PageTemplate(id="PT1", frames=[frame1, frame2, frame3])

        doc.addPageTemplates([page_template])

        story = [p, p_2, p_3]

        doc.build(story)


class SettingsDialog(QDialog):
    def __init__(self, parent=None):
        super(SettingsDialog, self).__init__(parent)
        self.ui = Ui_Dialog()
        self.ui.setupUi(self)
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)
        self.mw = parent
        self.fill_path()
        self.fill_check()
        self.fill_function()

    def fill_function(self):
        self.ui.buttonBox.accepted.connect(self.save_settings)
        self.ui.buttonBox.rejected.connect(self.close_dialog)
        self.ui.toolButton.clicked.connect(lambda: self.browse(1))
        self.ui.toolButton_2.clicked.connect(lambda: self.browse(2))
        self.ui.toolButton_3.clicked.connect(lambda: self.browse(3))

    def get_mw(self):
        return self.mw

    @MainWindow.dec
    def fill_check(self):
        check = [
            self.ui.checkBox,
            self.ui.checkBox_2,
            self.ui.checkBox_3,
            self.ui.checkBox_4,
            self.ui.checkBox_5,
            self.ui.checkBox_6,
            self.ui.checkBox_7,
            self.ui.checkBox_8,
            self.ui.checkBox_9,
            self.ui.checkBox_10,
            self.ui.checkBox_11,
        ]
        for i, b, n in zip(check, self.mw.ws["1"], range(1, 12)):
            if self.mw.settings.value(f"CheckBoxes/{n}", False, bool):
                i.setChecked(True)
            else:
                i.setChecked(False)
            i.setText(b.value)
        self.ui.groupBox.setEnabled(True)

    def getcol(self, path):
        try:
            fn = path
            wb = op.load_workbook(fn)
            ws = wb.active
            check = [
                self.ui.checkBox,
                self.ui.checkBox_2,
                self.ui.checkBox_3,
                self.ui.checkBox_4,
                self.ui.checkBox_5,
                self.ui.checkBox_6,
                self.ui.checkBox_7,
                self.ui.checkBox_8,
                self.ui.checkBox_9,
                self.ui.checkBox_10,
                self.ui.checkBox_11,
            ]
            for i, b in zip(check, ws["1"]):
                i.setText(b.value)
                i.setChecked(True)
                wb.save(fn)
                wb.close()
            self.ui.groupBox.setEnabled(True)
        except Exception as e:
            print("getcol")
            self.error(e)

    def fill_path(self):
        try:
            self.ui.lineEdit.setText(self.mw.settings.value("Path/Path_1"))
            self.ui.lineEdit_2.setText(self.mw.settings.value("Path/Path_2"))
            self.ui.lineEdit_3.setText(self.mw.settings.value("Path/Path_3"))
            self.ui.plainTextEdit.setPlainText(self.mw.settings.value("Path/Branches"))
            self.ui.plainTextEdit_2.setPlainText(self.mw.settings.value("Path/Names"))
        except:
            pass

    def browse(self, n):
        if n == 1:
            prev = self.ui.lineEdit.text()
            file, _ = QFileDialog.getOpenFileName(self, "Open file", prev, "(*.xlsx)")
            if file:
                self.ui.lineEdit.setText(file)
                self.getcol(file)
        elif n == 2:
            prev = self.ui.lineEdit_2.text()
            file = QFileDialog.getExistingDirectory(self, "Save file", prev)
            if file:
                self.ui.lineEdit_2.setText(file)
        elif n == 3:
            prev = self.ui.lineEdit_3.text()
            file = QFileDialog.getExistingDirectory(self, "Save file", prev)
            if file:
                self.ui.lineEdit_3.setText(file)

    def save_settings(self):
        self.mw.settings.setValue("Path/Path_1", self.ui.lineEdit.text())
        self.mw.settings.setValue("Path/Path_2", self.ui.lineEdit_2.text())
        self.mw.settings.setValue("Path/Path_3", self.ui.lineEdit_3.text())
        self.mw.settings.setValue("Path/Branches", self.ui.plainTextEdit.toPlainText())
        self.mw.settings.setValue("Path/Names", self.ui.plainTextEdit_2.toPlainText())
        check = [
            self.ui.checkBox,
            self.ui.checkBox_2,
            self.ui.checkBox_3,
            self.ui.checkBox_4,
            self.ui.checkBox_5,
            self.ui.checkBox_6,
            self.ui.checkBox_7,
            self.ui.checkBox_8,
            self.ui.checkBox_9,
            self.ui.checkBox_10,
            self.ui.checkBox_11,
        ]

        self.mw.settings.remove("CheckBoxes")
        for i, n in zip(check, range(1, 12)):
            if i.isChecked():
                self.mw.settings.setValue(f"CheckBoxes/{n}", True)
            else:
                self.mw.settings.setValue(f"CheckBoxes/{n}", False)

        self.mw.fill_branches()
        self.mw.fill_table()
        self.close()

    def close_dialog(self):
        self.close()


class CustomCalendarWidget(QCalendarWidget):
    def __init__(self):
        super().__init__()
        self.ui = Ui_Calendar()
        self.ui.setupUi(self)
        self.setMinimumSize(281, 201)


class DropArea(QWidget):
    file_dragged = pyqtSignal()
    selection_changed = pyqtSignal(str, str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)
        self.list_widget = QListWidget(self)

        self.list_widget.itemSelectionChanged.connect(self.selected)
        layout = QVBoxLayout()
        layout.addWidget(self.list_widget)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        self.setLayout(layout)

    def selected(self):
        items_selected = self.list_widget.selectedItems()
        for item in items_selected:
            name = item.text()
            data = item.data(Qt.UserRole)
            self.selection_changed.emit(name, data)

    def dragEnterEvent(self, event: QDragEnterEvent):
        event.acceptProposedAction()

    def dropEvent(self, event: QDropEvent):
        for url in event.mimeData().urls():
            count = 0
            for i in range(self.list_widget.count()):
                item = self.list_widget.item(i)
                if item.data(Qt.UserRole) == url.toLocalFile():
                    count += 1
            if count == 0:
                list_item = QListWidgetItem(QIcon("assets/file.svg"), url.fileName())
                list_item.setData(Qt.UserRole, url.toLocalFile())
                self.list_widget.addItem(list_item)
                self.file_dragged.emit()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    mainWindow = MainWindow()
    mainWindow.show()
    sys.exit(app.exec_())
